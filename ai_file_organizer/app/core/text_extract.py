"""
Text extraction utilities for various file types.
Converts files to clean, readable text suitable for AI analysis.
"""

import csv
import logging
from pathlib import Path
from typing import Optional, List, Dict, Any
import io

logger = logging.getLogger(__name__)

# Maximum limits to prevent memory issues
MAX_ROWS_TO_READ = 100
MAX_CHARS = 8000
MAX_CELL_LENGTH = 200


def extract_csv_text(file_path: Path, max_rows: int = MAX_ROWS_TO_READ) -> Optional[str]:
    """
    Extract readable text from a CSV file.
    
    Creates a summary including:
    - Column headers
    - Sample rows (formatted as readable text)
    - Basic statistics (row count, column count)
    
    Args:
        file_path: Path to the CSV file
        max_rows: Maximum number of sample rows to include
        
    Returns:
        Formatted text summary of the CSV, or None on error
    """
    try:
        # Try different encodings
        content = None
        for encoding in ['utf-8', 'utf-8-sig', 'latin-1', 'cp1252']:
            try:
                with open(file_path, 'r', encoding=encoding, newline='') as f:
                    content = f.read(500000)  # Read up to 500KB
                break
            except UnicodeDecodeError:
                continue
        
        if content is None:
            logger.warning(f"Could not decode CSV file: {file_path}")
            return None
        
        # Parse CSV
        reader = csv.reader(io.StringIO(content))
        rows: List[List[str]] = []
        
        for i, row in enumerate(reader):
            if i >= max_rows + 1:  # +1 for header
                break
            # Truncate long cells
            row = [cell[:MAX_CELL_LENGTH] + '...' if len(cell) > MAX_CELL_LENGTH else cell for cell in row]
            rows.append(row)
        
        if not rows:
            return None
        
        # Extract headers and data
        headers = rows[0] if rows else []
        data_rows = rows[1:] if len(rows) > 1 else []
        
        # Count total rows (approximate for large files)
        total_rows = content.count('\n')
        
        # Build readable summary
        parts: List[str] = []
        
        # Header info
        parts.append(f"CSV File: {file_path.name}")
        parts.append(f"Columns ({len(headers)}): {', '.join(headers[:20])}")
        if len(headers) > 20:
            parts.append(f"  ... and {len(headers) - 20} more columns")
        parts.append(f"Approximate rows: {total_rows}")
        parts.append("")
        
        # Sample data as readable text
        parts.append("Sample Data:")
        for i, row in enumerate(data_rows[:15]):  # Show up to 15 sample rows
            if headers and len(row) == len(headers):
                # Format as "Header: Value" pairs
                pairs = [f"{h}: {v}" for h, v in zip(headers, row) if v.strip()]
                if pairs:
                    parts.append(f"  Row {i+1}: {'; '.join(pairs[:8])}")
                    if len(pairs) > 8:
                        parts.append(f"    ... ({len(pairs) - 8} more fields)")
            else:
                # Just list values
                parts.append(f"  Row {i+1}: {', '.join(row[:10])}")
        
        if len(data_rows) > 15:
            parts.append(f"  ... ({len(data_rows) - 15} more sample rows)")
        
        # Add column analysis
        if headers and data_rows:
            parts.append("")
            parts.append("Column Analysis:")
            for j, header in enumerate(headers[:10]):
                # Get unique values from this column
                col_values = [row[j] for row in data_rows if j < len(row) and row[j].strip()]
                unique_count = len(set(col_values))
                sample_values = list(set(col_values))[:5]
                if sample_values:
                    parts.append(f"  {header}: {unique_count} unique values. Examples: {', '.join(sample_values)}")
        
        result = '\n'.join(parts)
        
        # Ensure we don't exceed max chars
        if len(result) > MAX_CHARS:
            result = result[:MAX_CHARS] + "\n... (truncated)"
        
        logger.info(f"Extracted {len(result)} chars from CSV: {file_path.name}")
        return result
        
    except Exception as e:
        logger.error(f"Error extracting CSV text from {file_path}: {e}")
        return None


def extract_text_file_content(file_path: Path, max_chars: int = MAX_CHARS) -> Optional[str]:
    """
    Extract text content from plain text files (.txt, .md, .json, .xml, etc).
    
    Args:
        file_path: Path to the text file
        max_chars: Maximum characters to extract
        
    Returns:
        Text content or None on error
    """
    try:
        for encoding in ['utf-8', 'utf-8-sig', 'latin-1', 'cp1252']:
            try:
                with open(file_path, 'r', encoding=encoding, errors='ignore') as f:
                    content = f.read(max_chars)
                return content
            except UnicodeDecodeError:
                continue
        return None
    except Exception as e:
        logger.error(f"Error reading text file {file_path}: {e}")
        return None


def extract_spreadsheet_text(file_path: Path) -> Optional[str]:
    """
    Extract text from Excel files (.xlsx, .xls).
    Requires openpyxl for .xlsx files.
    
    Args:
        file_path: Path to the spreadsheet
        
    Returns:
        Formatted text summary or None
    """
    ext = file_path.suffix.lower()
    
    if ext == '.xlsx':
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            
            parts: List[str] = []
            parts.append(f"Excel File: {file_path.name}")
            parts.append(f"Sheets: {', '.join(wb.sheetnames)}")
            parts.append("")
            
            for sheet_name in wb.sheetnames[:3]:  # First 3 sheets
                sheet = wb[sheet_name]
                parts.append(f"Sheet: {sheet_name}")
                
                rows_data: List[List[str]] = []
                for i, row in enumerate(sheet.iter_rows(max_row=20, values_only=True)):
                    if i >= 20:
                        break
                    row_text = [str(cell)[:100] if cell is not None else '' for cell in row[:15]]
                    rows_data.append(row_text)
                
                if rows_data:
                    headers = rows_data[0]
                    parts.append(f"  Columns: {', '.join([h for h in headers if h])}")
                    for j, row in enumerate(rows_data[1:10]):
                        if headers:
                            pairs = [f"{h}: {v}" for h, v in zip(headers, row) if v]
                            if pairs:
                                parts.append(f"    Row {j+1}: {'; '.join(pairs[:6])}")
                parts.append("")
            
            wb.close()
            
            result = '\n'.join(parts)
            if len(result) > MAX_CHARS:
                result = result[:MAX_CHARS] + "\n... (truncated)"
            return result
            
        except ImportError:
            logger.warning("openpyxl not installed - cannot read .xlsx files")
            return None
        except Exception as e:
            logger.error(f"Error reading Excel file {file_path}: {e}")
            return None
    
    return None


def extract_pdf_text(file_path: Path, max_pages: int = 5) -> Optional[str]:
    """
    Extract text from PDF files using PyPDF2.
    
    This is a pure Python solution that doesn't require poppler.
    
    Args:
        file_path: Path to the PDF file
        max_pages: Maximum number of pages to extract
        
    Returns:
        Extracted text content or None
    """
    try:
        from PyPDF2 import PdfReader
        
        reader = PdfReader(str(file_path))
        parts: List[str] = []
        
        parts.append(f"PDF File: {file_path.name}")
        parts.append(f"Total Pages: {len(reader.pages)}")
        parts.append("")
        
        for i, page in enumerate(reader.pages[:max_pages]):
            try:
                text = page.extract_text()
                if text and text.strip():
                    # Clean up the text
                    text = ' '.join(text.split())  # Normalize whitespace
                    if len(text) > 2000:
                        text = text[:2000] + "..."
                    parts.append(f"Page {i+1}:")
                    parts.append(text)
                    parts.append("")
            except Exception as e:
                logger.debug(f"Could not extract text from page {i+1}: {e}")
                continue
        
        if len(reader.pages) > max_pages:
            parts.append(f"... ({len(reader.pages) - max_pages} more pages)")
        
        result = '\n'.join(parts)
        
        if len(result) > MAX_CHARS:
            result = result[:MAX_CHARS] + "\n... (truncated)"
        
        # Only return if we got meaningful content
        if len(result) > 100:
            logger.info(f"Extracted {len(result)} chars from PDF: {file_path.name}")
            return result
        return None
        
    except ImportError:
        logger.warning("PyPDF2 not installed - cannot read PDF files")
        return None
    except Exception as e:
        logger.error(f"Error extracting PDF text from {file_path}: {e}")
        return None


def extract_file_text(file_path: Path) -> Optional[str]:
    """
    Extract readable text from a file based on its type.
    
    Dispatcher function that chooses the appropriate extractor.
    
    Args:
        file_path: Path to the file
        
    Returns:
        Extracted text content or None
    """
    if not file_path.exists():
        return None
    
    ext = file_path.suffix.lower()
    
    # CSV files
    if ext == '.csv':
        return extract_csv_text(file_path)
    
    # Excel files
    if ext in {'.xlsx', '.xls'}:
        return extract_spreadsheet_text(file_path)
    
    # PDF files
    if ext == '.pdf':
        return extract_pdf_text(file_path)
    
    # Plain text files
    text_extensions = {
        '.txt', '.md', '.json', '.xml', '.html', '.htm',
        '.log', '.ini', '.cfg', '.yaml', '.yml', '.toml',
        '.py', '.js', '.ts', '.jsx', '.tsx', '.css', '.scss',
        '.java', '.cpp', '.c', '.h', '.hpp', '.cs', '.go',
        '.rb', '.php', '.swift', '.kt', '.rs', '.sql',
        '.sh', '.bat', '.ps1', '.env', '.gitignore'
    }
    if ext in text_extensions:
        return extract_text_file_content(file_path)
    
    # For other files, try to read as text
    try:
        return extract_text_file_content(file_path, max_chars=4000)
    except Exception:
        return None


def get_supported_text_formats() -> List[str]:
    """Return list of file extensions that can be text-extracted."""
    return [
        '.csv', '.xlsx', '.xls',
        '.txt', '.md', '.json', '.xml', '.html', '.htm',
        '.log', '.ini', '.cfg', '.yaml', '.yml', '.toml',
        '.py', '.js', '.ts', '.jsx', '.tsx', '.css', '.scss',
        '.java', '.cpp', '.c', '.h', '.hpp', '.cs', '.go',
        '.rb', '.php', '.swift', '.kt', '.rs', '.sql',
        '.sh', '.bat', '.ps1'
    ]
