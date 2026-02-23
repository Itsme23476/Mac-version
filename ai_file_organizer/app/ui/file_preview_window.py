"""
File Preview Window - A floating preview window that appears over fullscreen apps.
Supports images, videos (with full controls), PDFs, text files, and Excel files.
Uses the same macOS window management techniques as QuickSearchOverlay.
"""

from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, 
    QFrame, QGraphicsDropShadowEffect, QPlainTextEdit, QWidget,
    QSizePolicy, QScrollArea, QSlider, QTableWidget, QTableWidgetItem,
    QHeaderView, QStyle
)
from PySide6.QtCore import Qt, QTimer, QUrl, QSize, Signal
from PySide6.QtGui import QPixmap, QColor, QFont, QMovie, QGuiApplication
import logging
import sys
import os

logger = logging.getLogger(__name__)

# File type categories
IMAGE_EXTENSIONS = {'.jpg', '.jpeg', '.png', '.gif', '.webp', '.bmp', '.tiff', '.tif', '.ico', '.heic', '.heif'}
VIDEO_EXTENSIONS = {'.mp4', '.mov', '.webm', '.avi', '.mkv', '.m4v', '.wmv', '.flv', '.mpeg', '.mpg', '.3gp'}
PDF_EXTENSIONS = {'.pdf'}
EXCEL_EXTENSIONS = {'.xlsx', '.xls', '.xlsm', '.xlsb', '.csv'}
TEXT_EXTENSIONS = {'.txt', '.md', '.json', '.py', '.js', '.ts', '.html', '.css', '.xml', '.yaml', '.yml', 
                   '.ini', '.cfg', '.conf', '.log', '.sh', '.bash', '.zsh', '.fish', '.swift', '.kt',
                   '.java', '.c', '.cpp', '.h', '.hpp', '.rs', '.go', '.rb', '.php', '.sql', '.r'}


def get_file_type(file_path: str) -> str:
    """Determine the type of file based on extension."""
    ext = os.path.splitext(file_path)[1].lower()
    if ext in IMAGE_EXTENSIONS:
        return 'image'
    elif ext in VIDEO_EXTENSIONS:
        return 'video'
    elif ext in PDF_EXTENSIONS:
        return 'pdf'
    elif ext in EXCEL_EXTENSIONS:
        return 'excel'
    elif ext in TEXT_EXTENSIONS:
        return 'text'
    else:
        return 'unknown'


def format_time(ms: int) -> str:
    """Format milliseconds as mm:ss or hh:mm:ss."""
    if ms < 0:
        ms = 0
    seconds = ms // 1000
    minutes = seconds // 60
    hours = minutes // 60
    seconds = seconds % 60
    minutes = minutes % 60
    
    if hours > 0:
        return f"{hours}:{minutes:02d}:{seconds:02d}"
    return f"{minutes}:{seconds:02d}"


class VideoControlBar(QWidget):
    """Custom video control bar with play/pause, timeline, volume, and timestamps."""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self._media_player = None
        self._is_seeking = False
        
        self._build_ui()
    
    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 8, 12, 12)
        layout.setSpacing(8)
        
        # Timeline slider (progress bar)
        self.timeline = QSlider(Qt.Horizontal)
        self.timeline.setObjectName("videoTimeline")
        self.timeline.setMinimum(0)
        self.timeline.setMaximum(1000)
        self.timeline.setValue(0)
        self.timeline.sliderPressed.connect(self._on_slider_pressed)
        self.timeline.sliderReleased.connect(self._on_slider_released)
        self.timeline.sliderMoved.connect(self._on_slider_moved)
        layout.addWidget(self.timeline)
        
        # Controls row
        controls_row = QHBoxLayout()
        controls_row.setSpacing(12)
        
        # Play/Pause button
        self.btn_play = QPushButton("▶")
        self.btn_play.setObjectName("videoPlayBtn")
        self.btn_play.setFixedSize(36, 36)
        self.btn_play.setCursor(Qt.PointingHandCursor)
        self.btn_play.clicked.connect(self._toggle_play)
        controls_row.addWidget(self.btn_play)
        
        # Current time label
        self.time_current = QLabel("0:00")
        self.time_current.setObjectName("videoTimeLabel")
        self.time_current.setFixedWidth(50)
        controls_row.addWidget(self.time_current)
        
        # Separator
        separator = QLabel("/")
        separator.setStyleSheet("color: #666;")
        controls_row.addWidget(separator)
        
        # Duration label
        self.time_duration = QLabel("0:00")
        self.time_duration.setObjectName("videoTimeLabel")
        self.time_duration.setFixedWidth(50)
        controls_row.addWidget(self.time_duration)
        
        controls_row.addStretch()
        
        # Volume icon
        self.btn_mute = QPushButton("🔊")
        self.btn_mute.setObjectName("videoMuteBtn")
        self.btn_mute.setFixedSize(32, 32)
        self.btn_mute.setCursor(Qt.PointingHandCursor)
        self.btn_mute.clicked.connect(self._toggle_mute)
        controls_row.addWidget(self.btn_mute)
        
        # Volume slider
        self.volume_slider = QSlider(Qt.Horizontal)
        self.volume_slider.setObjectName("videoVolume")
        self.volume_slider.setFixedWidth(80)
        self.volume_slider.setMinimum(0)
        self.volume_slider.setMaximum(100)
        self.volume_slider.setValue(100)
        self.volume_slider.valueChanged.connect(self._on_volume_changed)
        controls_row.addWidget(self.volume_slider)
        
        layout.addLayout(controls_row)
        
        self._apply_stylesheet()
    
    def _apply_stylesheet(self):
        self.setStyleSheet("""
            #videoTimeline {
                height: 6px;
            }
            #videoTimeline::groove:horizontal {
                background: #333;
                height: 6px;
                border-radius: 3px;
            }
            #videoTimeline::handle:horizontal {
                background: #7C4DFF;
                width: 14px;
                height: 14px;
                margin: -4px 0;
                border-radius: 7px;
            }
            #videoTimeline::sub-page:horizontal {
                background: #7C4DFF;
                border-radius: 3px;
            }
            #videoPlayBtn {
                background-color: #7C4DFF;
                border: none;
                border-radius: 18px;
                color: white;
                font-size: 14px;
            }
            #videoPlayBtn:hover {
                background-color: #9575FF;
            }
            #videoTimeLabel {
                color: #aaa;
                font-size: 12px;
                font-family: 'SF Mono', 'Menlo', monospace;
            }
            #videoMuteBtn {
                background: transparent;
                border: none;
                font-size: 16px;
            }
            #videoMuteBtn:hover {
                background-color: #333;
                border-radius: 4px;
            }
            #videoVolume {
                height: 4px;
            }
            #videoVolume::groove:horizontal {
                background: #333;
                height: 4px;
                border-radius: 2px;
            }
            #videoVolume::handle:horizontal {
                background: #aaa;
                width: 12px;
                height: 12px;
                margin: -4px 0;
                border-radius: 6px;
            }
            #videoVolume::sub-page:horizontal {
                background: #666;
                border-radius: 2px;
            }
        """)
    
    def set_media_player(self, player, audio_output):
        """Connect to a QMediaPlayer instance."""
        self._media_player = player
        self._audio_output = audio_output
        
        if player:
            player.positionChanged.connect(self._on_position_changed)
            player.durationChanged.connect(self._on_duration_changed)
            player.playbackStateChanged.connect(self._on_playback_state_changed)
    
    def _on_position_changed(self, position):
        """Update timeline and current time label."""
        if not self._is_seeking and self._media_player:
            duration = self._media_player.duration()
            if duration > 0:
                self.timeline.setValue(int(position * 1000 / duration))
            self.time_current.setText(format_time(position))
    
    def _on_duration_changed(self, duration):
        """Update duration label."""
        self.time_duration.setText(format_time(duration))
    
    def _on_playback_state_changed(self, state):
        """Update play button icon based on state."""
        from PySide6.QtMultimedia import QMediaPlayer
        if state == QMediaPlayer.PlayingState:
            self.btn_play.setText("⏸")
        else:
            self.btn_play.setText("▶")
    
    def _toggle_play(self):
        """Toggle play/pause."""
        if self._media_player:
            from PySide6.QtMultimedia import QMediaPlayer
            if self._media_player.playbackState() == QMediaPlayer.PlayingState:
                self._media_player.pause()
            else:
                self._media_player.play()
    
    def _on_slider_pressed(self):
        """User started seeking."""
        self._is_seeking = True
    
    def _on_slider_released(self):
        """User finished seeking."""
        if self._media_player:
            duration = self._media_player.duration()
            position = int(self.timeline.value() * duration / 1000)
            self._media_player.setPosition(position)
        self._is_seeking = False
    
    def _on_slider_moved(self, value):
        """Update time label while seeking."""
        if self._media_player:
            duration = self._media_player.duration()
            position = int(value * duration / 1000)
            self.time_current.setText(format_time(position))
    
    def _toggle_mute(self):
        """Toggle mute."""
        if self._audio_output:
            is_muted = self._audio_output.isMuted()
            self._audio_output.setMuted(not is_muted)
            self.btn_mute.setText("🔇" if not is_muted else "🔊")
    
    def _on_volume_changed(self, value):
        """Update volume."""
        if self._audio_output:
            self._audio_output.setVolume(value / 100.0)
            if value == 0:
                self.btn_mute.setText("🔇")
            else:
                self.btn_mute.setText("🔊")
    
    def reset(self):
        """Reset controls to initial state."""
        self.timeline.setValue(0)
        self.time_current.setText("0:00")
        self.time_duration.setText("0:00")
        self.btn_play.setText("▶")


class FilePreviewWindow(QDialog):
    """
    A floating preview window that can display images, videos, PDFs, Excel, and text files.
    Uses the same macOS window management as QuickSearchOverlay to appear over fullscreen apps.
    """
    
    def __init__(self, parent=None):
        super().__init__(parent)
        
        self.setWindowTitle("File Preview")
        self.setWindowFlags(
            Qt.Window |
            Qt.WindowStaysOnTopHint |
            Qt.FramelessWindowHint |
            Qt.Tool
        )
        self.setAttribute(Qt.WA_TranslucentBackground)
        
        # Initial size - will be adjusted by position_near_popup based on available space
        self.resize(600, 500)
        
        # Track current file
        self._current_file = None
        self._movie = None  # For animated GIFs
        self._media_player = None
        self._audio_output = None
        self._video_widget = None
        
        # macOS configuration flags
        self._macos_panel_configured = False
        self._level_timer = None
        self._drag_pos = None
        
        # Build UI
        self._build_ui()
        
        # Setup macOS-specific window management
        if sys.platform == 'darwin':
            self._setup_macos_timer()
    
    def _build_ui(self):
        """Build the preview window UI."""
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(10, 10, 10, 10)
        
        # Container frame with shadow
        self.container = QFrame()
        self.container.setObjectName("previewFrame")
        
        shadow = QGraphicsDropShadowEffect(self)
        shadow.setBlurRadius(25)
        shadow.setColor(QColor(0, 0, 0, 180))
        shadow.setOffset(0, 8)
        self.container.setGraphicsEffect(shadow)
        
        main_layout.addWidget(self.container)
        
        # Layout inside container
        layout = QVBoxLayout(self.container)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)
        
        # Header with title and close button
        header = QFrame()
        header.setObjectName("previewHeader")
        header.setFixedHeight(40)
        header_layout = QHBoxLayout(header)
        header_layout.setContentsMargins(16, 8, 8, 8)
        
        self.title_label = QLabel("Preview")
        self.title_label.setObjectName("previewTitle")
        header_layout.addWidget(self.title_label, 1)
        
        # Open externally button
        self.btn_external = QPushButton("↗")
        self.btn_external.setObjectName("previewExternalBtn")
        self.btn_external.setFixedSize(28, 28)
        self.btn_external.setCursor(Qt.PointingHandCursor)
        self.btn_external.setToolTip("Open in default app")
        self.btn_external.clicked.connect(self._open_externally)
        header_layout.addWidget(self.btn_external)
        
        self.close_btn = QPushButton("✕")
        self.close_btn.setObjectName("previewCloseBtn")
        self.close_btn.setFixedSize(24, 24)
        self.close_btn.setCursor(Qt.PointingHandCursor)
        self.close_btn.setToolTip("Close")
        self.close_btn.clicked.connect(self.hide)
        header_layout.addWidget(self.close_btn)
        
        layout.addWidget(header)
        
        # Content area - will hold the preview
        self.content_area = QFrame()
        self.content_area.setObjectName("previewContent")
        self.content_layout = QVBoxLayout(self.content_area)
        self.content_layout.setContentsMargins(0, 0, 0, 0)
        self.content_layout.setSpacing(0)
        
        # Placeholder label
        self.placeholder = QLabel("Select a file to preview")
        self.placeholder.setAlignment(Qt.AlignCenter)
        self.placeholder.setStyleSheet("color: #888; font-size: 14px;")
        self.content_layout.addWidget(self.placeholder)
        
        # Image label (hidden by default)
        self.image_label = QLabel()
        self.image_label.setAlignment(Qt.AlignCenter)
        self.image_label.setContentsMargins(16, 16, 16, 16)
        self.image_label.hide()
        self.content_layout.addWidget(self.image_label)
        
        # Text viewer (hidden by default)
        self.text_viewer = QPlainTextEdit()
        self.text_viewer.setReadOnly(True)
        self.text_viewer.setObjectName("previewTextViewer")
        self.text_viewer.hide()
        self.content_layout.addWidget(self.text_viewer)
        
        # Video container (hidden by default)
        self.video_container = QFrame()
        self.video_container.setObjectName("previewVideoContainer")
        self.video_container.hide()
        video_layout = QVBoxLayout(self.video_container)
        video_layout.setContentsMargins(0, 0, 0, 0)
        video_layout.setSpacing(0)
        self.content_layout.addWidget(self.video_container)
        
        # Excel table viewer (hidden by default)
        self.excel_table = QTableWidget()
        self.excel_table.setObjectName("previewExcelTable")
        self.excel_table.hide()
        self.content_layout.addWidget(self.excel_table)
        
        # PDF container (hidden by default)
        self.pdf_container = QFrame()
        self.pdf_container.setObjectName("previewPdfContainer")
        self.pdf_container.hide()
        pdf_layout = QVBoxLayout(self.pdf_container)
        pdf_layout.setContentsMargins(0, 0, 0, 0)
        self.content_layout.addWidget(self.pdf_container)
        
        # Info label for unsupported files
        self.info_label = QLabel()
        self.info_label.setAlignment(Qt.AlignCenter)
        self.info_label.setWordWrap(True)
        self.info_label.setContentsMargins(16, 16, 16, 16)
        self.info_label.hide()
        self.content_layout.addWidget(self.info_label)
        
        layout.addWidget(self.content_area, 1)
        
        # Video control bar (hidden by default, shown for videos)
        self.video_controls = VideoControlBar()
        self.video_controls.setObjectName("videoControlBar")
        self.video_controls.hide()
        layout.addWidget(self.video_controls)
        
        # Footer with file path
        footer = QFrame()
        footer.setObjectName("previewFooter")
        footer.setFixedHeight(32)
        footer_layout = QHBoxLayout(footer)
        footer_layout.setContentsMargins(16, 6, 16, 6)
        
        self.path_label = QLabel("")
        self.path_label.setObjectName("previewPath")
        footer_layout.addWidget(self.path_label, 1)
        
        layout.addWidget(footer)
        
        # Apply stylesheet
        self._apply_stylesheet()
    
    def _apply_stylesheet(self):
        """Apply dark theme stylesheet."""
        self.setStyleSheet("""
            #previewFrame {
                background-color: #1e1e1e;
                border-radius: 12px;
                border: 1px solid #333;
            }
            #previewHeader {
                background-color: #252525;
                border-top-left-radius: 12px;
                border-top-right-radius: 12px;
                border-bottom: 1px solid #333;
            }
            #previewTitle {
                color: #ffffff;
                font-size: 13px;
                font-weight: 500;
            }
            #previewCloseBtn {
                background-color: #ff5f57;
                border: none;
                border-radius: 12px;
                color: white;
                font-size: 14px;
                font-weight: bold;
            }
            #previewCloseBtn:hover {
                background-color: #ff3b30;
            }
            #previewExternalBtn {
                background-color: transparent;
                border: none;
                border-radius: 14px;
                color: #888;
                font-size: 18px;
                font-weight: bold;
            }
            #previewExternalBtn:hover {
                background-color: #444;
                color: white;
            }
            #previewContent {
                background-color: #1e1e1e;
            }
            #previewTextViewer {
                background-color: #252525;
                color: #e0e0e0;
                border: none;
                font-family: 'SF Mono', 'Menlo', 'Monaco', monospace;
                font-size: 12px;
                padding: 12px;
            }
            #previewFooter {
                background-color: #252525;
                border-bottom-left-radius: 12px;
                border-bottom-right-radius: 12px;
                border-top: 1px solid #333;
            }
            #previewPath {
                color: #666;
                font-size: 11px;
            }
            #previewVideoContainer {
                background-color: #000;
            }
            #videoControlBar {
                background-color: #1a1a1a;
                border-top: 1px solid #333;
            }
            #previewExcelTable {
                background-color: #252525;
                color: #e0e0e0;
                border: none;
                gridline-color: #333;
                font-size: 12px;
            }
            #previewExcelTable::item {
                padding: 4px 8px;
            }
            #previewExcelTable QHeaderView::section {
                background-color: #333;
                color: #aaa;
                border: none;
                padding: 6px;
                font-weight: bold;
            }
        """)
    
    def _open_externally(self):
        """Open the current file in the default application."""
        if self._current_file and os.path.exists(self._current_file):
            from PySide6.QtGui import QDesktopServices
            QDesktopServices.openUrl(QUrl.fromLocalFile(self._current_file))
    
    def preview_file(self, file_path: str):
        """
        Preview the specified file.
        
        Args:
            file_path: Path to the file to preview
        """
        if not file_path or not os.path.exists(file_path):
            self._show_error(f"File not found: {file_path}")
            return
        
        self._current_file = file_path
        file_type = get_file_type(file_path)
        file_name = os.path.basename(file_path)
        
        # Update title and path
        self.title_label.setText(f"{file_name}")
        self.path_label.setText(file_path)
        
        # Hide all content widgets
        self._hide_all_content()
        
        # Stop any playing media
        self._stop_media()
        
        # Show appropriate preview based on file type
        if file_type == 'image':
            self._show_image(file_path)
        elif file_type == 'video':
            self._show_video(file_path)
        elif file_type == 'pdf':
            self._show_pdf(file_path)
        elif file_type == 'excel':
            self._show_excel(file_path)
        elif file_type == 'text':
            self._show_text(file_path)
        else:
            self._show_file_info(file_path)
        
        # Show the window
        self.show()
        self._bring_to_front()
    
    def _hide_all_content(self):
        """Hide all content widgets."""
        self.placeholder.hide()
        self.image_label.hide()
        self.text_viewer.hide()
        self.video_container.hide()
        self.video_controls.hide()
        self.excel_table.hide()
        self.pdf_container.hide()
        self.info_label.hide()
    
    def _stop_media(self):
        """Stop any playing media."""
        if self._movie:
            self._movie.stop()
            self._movie = None
        if self._media_player:
            self._media_player.stop()
        self.video_controls.reset()
    
    def _show_image(self, file_path: str):
        """Display an image file."""
        try:
            ext = os.path.splitext(file_path)[1].lower()
            
            # Handle animated GIFs
            if ext == '.gif':
                self._movie = QMovie(file_path)
                if self._movie.isValid():
                    # Scale the movie to fit
                    self._movie.setScaledSize(self._get_scaled_size(self._movie.frameRect().size()))
                    self.image_label.setMovie(self._movie)
                    self._movie.start()
                    self.image_label.show()
                    return
            
            # Regular images
            pixmap = QPixmap(file_path)
            if pixmap.isNull():
                self._show_error(f"Could not load image: {file_path}")
                return
            
            # Scale to fit content area
            scaled = pixmap.scaled(
                self.content_area.width() - 32,
                self.content_area.height() - 32,
                Qt.KeepAspectRatio,
                Qt.SmoothTransformation
            )
            self.image_label.setPixmap(scaled)
            self.image_label.show()
            
        except Exception as e:
            logger.error(f"Error loading image: {e}")
            self._show_error(f"Error loading image: {e}")
    
    def _get_scaled_size(self, original_size: QSize) -> QSize:
        """Calculate scaled size to fit content area while maintaining aspect ratio."""
        max_width = self.content_area.width() - 32
        max_height = self.content_area.height() - 32
        
        if original_size.width() <= max_width and original_size.height() <= max_height:
            return original_size
        
        scale_w = max_width / original_size.width()
        scale_h = max_height / original_size.height()
        scale = min(scale_w, scale_h)
        
        return QSize(int(original_size.width() * scale), int(original_size.height() * scale))
    
    def _show_video(self, file_path: str):
        """Display a video file with full controls."""
        try:
            from PySide6.QtMultimedia import QMediaPlayer, QAudioOutput
            from PySide6.QtMultimediaWidgets import QVideoWidget
            
            # Create video widget if needed
            if self._video_widget is None:
                self._video_widget = QVideoWidget()
                self._video_widget.setMinimumSize(400, 300)
                self.video_container.layout().addWidget(self._video_widget)
            
            # Create media player if needed
            if self._media_player is None:
                self._media_player = QMediaPlayer()
                self._audio_output = QAudioOutput()
                self._audio_output.setVolume(1.0)
                self._media_player.setAudioOutput(self._audio_output)
                self._media_player.setVideoOutput(self._video_widget)
                
                # Connect control bar
                self.video_controls.set_media_player(self._media_player, self._audio_output)
            
            # Load video
            self._media_player.setSource(QUrl.fromLocalFile(file_path))
            
            # Show video container and controls
            self.video_container.show()
            self.video_controls.show()
            
            # Auto-play
            self._media_player.play()
            
            logger.info(f"[Preview] Playing video: {file_path}")
            
        except ImportError as e:
            logger.warning(f"QtMultimedia not available: {e}")
            self._show_error("Video preview requires QtMultimedia.\nInstall with: pip install PySide6")
        except Exception as e:
            logger.error(f"Error playing video: {e}")
            self._show_error(f"Error playing video: {e}")
    
    def _show_pdf(self, file_path: str):
        """Display a PDF file."""
        try:
            # Try using QPdfDocument (Qt 6.4+)
            from PySide6.QtPdf import QPdfDocument
            from PySide6.QtPdfWidgets import QPdfView
            
            # Create PDF view if not exists
            if not hasattr(self, '_pdf_view'):
                self._pdf_view = QPdfView()
                self._pdf_document = QPdfDocument(self)
                self._pdf_view.setDocument(self._pdf_document)
                self._pdf_view.setPageMode(QPdfView.PageMode.MultiPage)
                self._pdf_view.setZoomMode(QPdfView.ZoomMode.FitToWidth)
                self.pdf_container.layout().addWidget(self._pdf_view)
            
            # Load PDF
            self._pdf_document.load(file_path)
            self.pdf_container.show()
            
            logger.info(f"[Preview] Showing PDF: {file_path}")
            
        except ImportError:
            logger.warning("QPdfView not available, showing file info")
            self._show_pdf_fallback(file_path)
        except Exception as e:
            logger.error(f"Error loading PDF: {e}")
            self._show_error(f"Error loading PDF: {e}")
    
    def _show_pdf_fallback(self, file_path: str):
        """Fallback PDF display - show file info."""
        file_size = os.path.getsize(file_path)
        size_str = self._format_file_size(file_size)
        
        self.info_label.setText(
            f"📄 PDF Document\n\n"
            f"Size: {size_str}\n\n"
            f"PDF preview requires PySide6 PDF support.\n"
            f"Click ↗ to open in default PDF viewer."
        )
        self.info_label.setStyleSheet("color: #aaa; font-size: 14px;")
        self.info_label.show()
    
    def _show_excel(self, file_path: str):
        """Display an Excel/CSV file in a table."""
        try:
            ext = os.path.splitext(file_path)[1].lower()
            
            if ext == '.csv':
                self._show_csv(file_path)
            else:
                self._show_xlsx(file_path)
                
        except Exception as e:
            logger.error(f"Error loading Excel file: {e}")
            self._show_error(f"Error loading Excel file: {e}")
    
    def _show_csv(self, file_path: str):
        """Display a CSV file."""
        import csv
        
        try:
            with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
                reader = csv.reader(f)
                rows = list(reader)
            
            if not rows:
                self._show_error("Empty CSV file")
                return
            
            # Limit rows for performance
            max_rows = 100
            if len(rows) > max_rows:
                rows = rows[:max_rows]
                truncated = True
            else:
                truncated = False
            
            # Setup table
            self.excel_table.setRowCount(len(rows) - 1)  # First row as header
            self.excel_table.setColumnCount(len(rows[0]) if rows else 0)
            
            # Set headers from first row
            if rows:
                self.excel_table.setHorizontalHeaderLabels(rows[0])
            
            # Fill data
            for i, row in enumerate(rows[1:]):
                for j, cell in enumerate(row):
                    if j < self.excel_table.columnCount():
                        item = QTableWidgetItem(str(cell))
                        self.excel_table.setItem(i, j, item)
            
            # Resize columns to content
            self.excel_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
            self.excel_table.show()
            
            if truncated:
                logger.info(f"[Preview] CSV truncated to {max_rows} rows")
                
        except Exception as e:
            logger.error(f"Error reading CSV: {e}")
            self._show_error(f"Error reading CSV: {e}")
    
    def _show_xlsx(self, file_path: str):
        """Display an Excel file using openpyxl."""
        try:
            import openpyxl
            
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            
            # Get dimensions
            max_row = min(ws.max_row or 1, 100)  # Limit to 100 rows
            max_col = ws.max_column or 1
            
            # Setup table
            self.excel_table.setRowCount(max_row)
            self.excel_table.setColumnCount(max_col)
            
            # Generate column headers (A, B, C, ...)
            headers = []
            for i in range(max_col):
                col_letter = ''
                n = i
                while n >= 0:
                    col_letter = chr(65 + (n % 26)) + col_letter
                    n = n // 26 - 1
                headers.append(col_letter)
            self.excel_table.setHorizontalHeaderLabels(headers)
            
            # Fill data
            for i, row in enumerate(ws.iter_rows(max_row=max_row)):
                for j, cell in enumerate(row):
                    value = cell.value if cell.value is not None else ''
                    item = QTableWidgetItem(str(value))
                    self.excel_table.setItem(i, j, item)
            
            wb.close()
            
            self.excel_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
            self.excel_table.show()
            
            logger.info(f"[Preview] Showing Excel: {file_path}")
            
        except ImportError:
            logger.warning("openpyxl not available")
            self._show_excel_fallback(file_path)
        except Exception as e:
            logger.error(f"Error reading Excel: {e}")
            self._show_error(f"Error reading Excel file: {e}")
    
    def _show_excel_fallback(self, file_path: str):
        """Fallback for Excel files when openpyxl is not available."""
        file_size = os.path.getsize(file_path)
        size_str = self._format_file_size(file_size)
        ext = os.path.splitext(file_path)[1].lower()
        
        self.info_label.setText(
            f"📊 Excel File ({ext})\n\n"
            f"Size: {size_str}\n\n"
            f"Excel preview requires openpyxl.\n"
            f"Install with: pip install openpyxl\n\n"
            f"Click ↗ to open in Excel."
        )
        self.info_label.setStyleSheet("color: #aaa; font-size: 14px;")
        self.info_label.show()
    
    def _show_text(self, file_path: str):
        """Display a text file."""
        try:
            # Limit file size for preview (1MB max)
            file_size = os.path.getsize(file_path)
            if file_size > 1024 * 1024:
                with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
                    content = f.read(1024 * 1024)
                content += "\n\n... (file truncated, too large for preview)"
            else:
                with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
                    content = f.read()
            
            self.text_viewer.setPlainText(content)
            self.text_viewer.show()
            
            logger.info(f"[Preview] Showing text file: {file_path}")
            
        except Exception as e:
            logger.error(f"Error reading text file: {e}")
            self._show_error(f"Error reading file: {e}")
    
    def _show_file_info(self, file_path: str):
        """Show file info for unsupported types."""
        try:
            file_name = os.path.basename(file_path)
            file_size = os.path.getsize(file_path)
            ext = os.path.splitext(file_path)[1].lower() or "No extension"
            
            size_str = self._format_file_size(file_size)
            
            self.info_label.setText(
                f"📁 {file_name}\n\n"
                f"Type: {ext}\n"
                f"Size: {size_str}\n\n"
                f"Preview not available for this file type.\n"
                f"Click ↗ to open with default application."
            )
            self.info_label.setStyleSheet("color: #aaa; font-size: 14px;")
            self.info_label.show()
            
        except Exception as e:
            self._show_error(f"Error reading file info: {e}")
    
    def _show_error(self, message: str):
        """Show an error message."""
        self.info_label.setText(f"⚠️ {message}")
        self.info_label.setStyleSheet("color: #ff6b6b; font-size: 14px;")
        self.info_label.show()
    
    def _format_file_size(self, size_bytes: int) -> str:
        """Format file size in human-readable format."""
        for unit in ['B', 'KB', 'MB', 'GB']:
            if size_bytes < 1024:
                return f"{size_bytes:.1f} {unit}"
            size_bytes /= 1024
        return f"{size_bytes:.1f} TB"
    
    # ========== macOS Window Management ==========
    
    def _setup_macos_timer(self):
        """Setup timer to enforce window level on macOS."""
        self._level_timer = QTimer(self)
        self._level_timer.timeout.connect(self._enforce_window_level)
        self._level_timer.start(100)
    
    def _enforce_window_level(self):
        """Continuously enforce window level to combat Qt resetting it."""
        if not self.isVisible():
            return
        
        try:
            from AppKit import NSApp
            
            for ns_window in NSApp.windows():
                try:
                    if ns_window.title() == "File Preview":
                        current_level = ns_window.level()
                        if current_level != 1000:
                            ns_window.setLevel_(1000)
                        
                        current_behavior = ns_window.collectionBehavior()
                        if current_behavior != 257:
                            ns_window.setCollectionBehavior_(257)
                        break
                except:
                    continue
        except Exception:
            pass
    
    def showEvent(self, event):
        """Handle show event - configure macOS window."""
        super().showEvent(event)
        if sys.platform == 'darwin':
            QTimer.singleShot(10, self._configure_macos_window)
            QTimer.singleShot(50, self._bring_to_front)
    
    def hideEvent(self, event):
        """Handle hide event - stop media playback."""
        self._stop_media()
        super().hideEvent(event)
    
    def _configure_macos_window(self):
        """Configure macOS window for fullscreen overlay capability."""
        if sys.platform != 'darwin':
            return
        
        try:
            from AppKit import NSApp
            
            for ns_window in NSApp.windows():
                try:
                    if ns_window.title() == "File Preview":
                        self._configure_macos_panel(ns_window)
                        break
                except:
                    continue
        except Exception as e:
            logger.error(f"[Preview] Error configuring macOS window: {e}")
    
    def _configure_macos_window_immediate(self):
        """Configure macOS window IMMEDIATELY after show - no timer delay."""
        if sys.platform != 'darwin':
            return
        
        try:
            from AppKit import NSApp
            
            # Find our window
            for ns_window in NSApp.windows():
                try:
                    if ns_window.title() == "File Preview":
                        # Set window level HIGH (above other windows)
                        ns_window.setLevel_(1000)
                        
                        # Collection behavior for all spaces
                        ns_window.setCollectionBehavior_(257)
                        
                        # Don't hide when app loses focus
                        if hasattr(ns_window, 'setHidesOnDeactivate_'):
                            ns_window.setHidesOnDeactivate_(False)
                        
                        # Floating panel
                        if hasattr(ns_window, 'setFloatingPanel_'):
                            ns_window.setFloatingPanel_(True)
                        
                        # Bring to front immediately
                        ns_window.orderFrontRegardless()
                        
                        logger.info(f"[Preview] macOS window configured immediately, level={ns_window.level()}")
                        break
                except Exception as e:
                    logger.error(f"[Preview] Error configuring window: {e}")
                    continue
        except Exception as e:
            logger.error(f"[Preview] Error in immediate macOS config: {e}")
    
    def _bring_to_front_fallback(self):
        """Fallback method to ensure window is visible."""
        try:
            if sys.platform == 'darwin':
                from AppKit import NSApp
                
                for ns_window in NSApp.windows():
                    try:
                        if ns_window.title() == "File Preview":
                            # Ensure level is still high
                            if ns_window.level() < 1000:
                                ns_window.setLevel_(1000)
                            
                            # Force to front
                            ns_window.orderFrontRegardless()
                            break
                    except:
                        continue
            
            # Qt fallback
            self.raise_()
            self.activateWindow()
            
        except Exception as e:
            logger.error(f"[Preview] Fallback bring to front error: {e}")
    
    def _configure_macos_panel(self, ns_window):
        """Configure NSWindow for fullscreen overlay capability."""
        try:
            # Collection behavior: CanJoinAllSpaces | FullScreenAuxiliary
            COLLECTION_BEHAVIOR = 257
            WINDOW_LEVEL = 1000
            
            ns_window.setCollectionBehavior_(COLLECTION_BEHAVIOR)
            ns_window.setLevel_(WINDOW_LEVEL)
            
            if hasattr(ns_window, 'setHidesOnDeactivate_'):
                ns_window.setHidesOnDeactivate_(False)
            
            # For preview, we want to prevent activation (no need for keyboard focus)
            try:
                if hasattr(ns_window, '_setPreventsActivation_'):
                    ns_window._setPreventsActivation_(True)
            except Exception:
                pass
            
            if hasattr(ns_window, 'setFloatingPanel_'):
                ns_window.setFloatingPanel_(True)
            
            logger.info("[Preview] macOS window configured for fullscreen overlay")
            
        except Exception as e:
            logger.error(f"[Preview] Error configuring macOS panel: {e}")
    
    def _bring_to_front(self):
        """Bring the preview window to front on macOS."""
        if sys.platform != 'darwin':
            self.raise_()
            self.activateWindow()
            return
        
        try:
            from AppKit import NSApp
            
            for ns_window in NSApp.windows():
                try:
                    if ns_window.title() == "File Preview":
                        # Ensure high window level
                        ns_window.setLevel_(1000)
                        ns_window.setCollectionBehavior_(257)
                        
                        # Try CGS API but don't fail if it doesn't work
                        try:
                            from app.ui.quick_search_overlay import move_window_to_active_space
                            window_number = ns_window.windowNumber()
                            move_window_to_active_space(window_number)
                        except Exception:
                            pass  # CGS API optional
                        
                        # Bring to front - this is the critical part
                        ns_window.orderFrontRegardless()
                        logger.info(f"[Preview] Window brought to front, level={ns_window.level()}")
                        break
                except Exception as e:
                    logger.error(f"[Preview] Error with window: {e}")
                    continue
            
            self.raise_()
            
        except Exception as e:
            logger.error(f"[Preview] Error bringing window to front: {e}")
            self.raise_()
    
    def position_near_popup(self, popup_geometry):
        """Position the preview window next to the search popup."""
        screen = QGuiApplication.primaryScreen().geometry()
        
        if popup_geometry is None:
            # Center on screen
            self.resize(600, 500)
            self.move(
                (screen.width() - self.width()) // 2,
                (screen.height() - self.height()) // 2
            )
            return
        
        # Simple positioning: to the RIGHT of popup, aligned with top
        x = popup_geometry.right() + 20
        y = popup_geometry.top()
        
        # If it goes off screen to the right, position to the left instead
        if x + self.width() > screen.width():
            x = popup_geometry.left() - self.width() - 20
        
        # If still off screen, just center it
        if x < 0:
            x = (screen.width() - self.width()) // 2
        
        # Ensure Y is on screen
        if y + self.height() > screen.height():
            y = screen.height() - self.height() - 50
        if y < 50:
            y = 50
        
        self.move(int(x), int(y))
        logger.info(f"[Preview] Positioned at ({x}, {y})")
    
    # ========== Dragging Support ==========
    
    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self._drag_pos = event.globalPosition().toPoint() - self.frameGeometry().topLeft()
            event.accept()
    
    def mouseMoveEvent(self, event):
        if event.buttons() == Qt.LeftButton and self._drag_pos:
            self.move(event.globalPosition().toPoint() - self._drag_pos)
            event.accept()
    
    def mouseReleaseEvent(self, event):
        self._drag_pos = None
        event.accept()
    
    def keyPressEvent(self, event):
        """Handle key press - Escape to close, Space to play/pause video."""
        if event.key() == Qt.Key_Escape:
            self.hide()
        elif event.key() == Qt.Key_Space:
            # Toggle play/pause if video is playing
            if self._media_player and self.video_controls.isVisible():
                self.video_controls._toggle_play()
        else:
            super().keyPressEvent(event)
